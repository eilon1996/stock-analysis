import yfinance as yf
import pandas as pd
import time
import numpy as np
from bs4 import BeautifulSoup
import requests
import openpyxl
import xlsxwriter
import pathlib

import calculation


class FinanceProduct:

    def __init__(self, symbol):

        self.symbol = symbol.upper()
        try:
            ticker = yf.Ticker(symbol)
            ###### get stock info ######
            info = ticker.info

        except:
            # some times its not working on the first time becouse of internet connection
            time.sleep(7)
            ticker = yf.Ticker(symbol)
            ###### get stock info ######
            info = ticker.info

        self.name = (info['shortName'] if len(info['shortName']) < len(info['longName']) else info['longName'])

        self.currency = info["currency"]
        self.pe_ratio = info["trailingPE"]
        self.market_cap = info["marketCap"]
        self.avg_volume = info["averageVolume"]
        self.current_price = info["previousClose"]

        ##### determine product TYPE and more... #####

        try:

            if info['quoteType'] is not None and info['quoteType'] == "ETF":
                self.product_type = "ETF"

                ###### get sectors ########
                url = "https://finance.yahoo.com/quote/" + self.symbol + "/holdings?p=" + self.symbol
                html_content = requests.get(url).text  # Make a GET request to fetch the raw HTML content
                soup = BeautifulSoup(html_content, "lxml")  # Parse the html content
                data = soup.find("section", attrs={"class": "Pb(20px) smartphone_Px(20px) smartphone_Mt(20px)"})
                sectors = {}

                bond_share = calculation.convert_string_to_number(
                    data.contents[0].contents[0].contents[1].contents[1].contents[1].text[:-1])

                if bond_share >= 50:
                    self.sector = "Bonds etf"
                    sectors_data = data.contents[1].contents[1].contents[1].contents
                else:
                    self.sector = "Stocks etf"
                    sectors_data = data.contents[0].contents[1].contents[1].contents

                sectors = [calculation.convert_string_to_number(x.contents[2].text[:-1]) for x in sectors_data[1:]]
                max1 = max(sectors)
                max1_index = sectors.index(max1)
                self.industry = calculation.get_sectors_name(max1_index) + " " + str(max1) + "%"

                if max1 < 90:
                    sectors[max1_index] = 0
                    max2 = max(sectors)
                    max2_index = sectors.index(max2)
                    self.industry += "\n" + calculation.get_sectors_name(max2_index) + " " + str(max2) + "%"

                # ETF dont have profitability or debt_to_assent
                self.profitability = -1
                self.debt_to_assent = -1

            else:
                if info['quoteType'] is not None:
                    self.product_type = info['quoteType']
                else:
                    self.product_type = "Stock"

                self.sector = info["sector"]
                self.industry = info["industry"]

                ###### get stock profitability ######

                url = "https://finance.yahoo.com/quote/" + self.symbol + "/financials?p=" + self.symbol
                html_content = requests.get(url).text  # Make a GET request to fetch the raw HTML content
                soup = BeautifulSoup(html_content, "lxml")  # Parse the html content
                data = soup.find("div", attrs={"class": "D(tbrg)"})

                revenue = calculation.convert_string_to_number(data.contents[0].contents[0].contents[2].string)
                income = calculation.convert_string_to_number(data.contents[10].contents[0].contents[2].string)
                self.profitability = calculation.two_point_percentage(revenue / income)

                ###### get stock debt/assent ######

                url = "https://finance.yahoo.com/quote/" + self.symbol + "/balance-sheet?p=" + self.symbol
                html_content = requests.get(url).text  # Make a GET request to fetch the raw HTML content
                soup = BeautifulSoup(html_content, "lxml")  # Parse the html content
                data = soup.find("div", attrs={"class": "D(tbrg)"})

                assent = calculation.convert_string_to_number(data.contents[0].contents[0].contents[1].string)
                debt = calculation.convert_string_to_number(data.contents[11].contents[0].contents[1].string)
                self.debt_to_assent = calculation.two_point_percentage(debt / assent)


        except Exception as e:
            print("FP.init: ETF/Stock")
            print(e)

        ###### get stock Leveraged ######

        if info['category'] is not None:
            try:
                "Leveraged" in info['category']
                self.leveraged = True
            except:
                self.leveraged = False
        else:
            self.leveraged = False

        ###### get historical market data ######
        try:
            history = ticker.history(
                period="5y",
                interval="3mo",
                group_by='ticker')
            self.price_history = history.Open.values[:-1]  # the last one is today price
            not_nan = self.price_history == self.price_history
            self.price_history = self.price_history[not_nan]

            self.dates = np.array([[history.index.day[i], history.index.month[i], history.index.year[i]]
                                   for i in range(len(history.index) - 1)])
            self.dates = self.dates[not_nan]

        except Exception as e:
            print("FP.init: extract dividends in %")
            print(e)

        ###### extract dividends ######
        try:
            dividends = ticker.dividends
            dividends, dates = dividends.values, list(dividends.index.year)
            first_year = dates[-1] - 4
            start_index = dates.index(first_year)

            self.yearly_dividend = np.zeros(5)

            count_last_year = 0
            count_previes_year = 0
            for (div, y) in zip(dividends[start_index:], dates[start_index:]):
                self.yearly_dividend[y - first_year] += div

                if y - first_year == 3: count_previes_year += 1
                if y - first_year == 4: count_last_year += 1
            if count_previes_year == count_last_year:
                # this is a full year dividend
                astimate_last_div = self.yearly_dividend[-1]
                self.last_full_year_div = sum(self.yearly_dividend[-4:-1])
            else:
                # this is not a full year dividend
                ### estimate for full year
                ### we asume that there will be the same amount of dividend as last year
                ### and that the rimaining dividens are like the last one
                astimate_last_div = self.yearly_dividend[-1] + dividends[-1] * (count_previes_year - count_last_year)
                self.last_full_year_div = sum(self.yearly_dividend[-4-count_last_year:-1-count_last_year])



        except Exception as e:
            print("FP.init: extract dividends")
            print(e)

        ###### extract dividends -- in % to price ######
        try:
            dates = np.asarray([history.index.year[i] for i in range(len(history.index) - 1)])
            dates = dates[not_nan]
            start_index = np.where(dates == first_year)[0][0]

            sum_price = np.zeros(5)

            count_last_year = 0
            for (p, y) in zip(self.price_history[start_index:], dates[start_index:]):
                sum_price[y - first_year] += p

                if y - first_year == 4: count_last_year += 1

            last_full_year_sum_price = sum_price[-1] + self.price_history[-1] * (4 - count_previes_year)

            self.div_per_price = self.yearly_dividend[:-1] / (sum_price[:-1] / 4)
            self.div_per_price += astimate_last_div / (last_full_year_sum_price / 4)
            self.div_per_price = calculation.two_point_percentage(self.div_per_price)

        except Exception as e:
            print("FP.init: extract dividends in %")
            print(e)

        ##### calculate div and price yield #####
        try:
            self.yield_1y = calculation.two_point_percentage(self.price_history[-1] / self.price_history[-5] - 1)
            if len(self.price_history) >= 21:
                # we have full data of 5 years
                self.yield_5y = self.price_history[-1] / self.price_history[0] - 1
            elif len(self.price_history) < 17:
                # not enough details to astimate for 5 years
                self.yield_5y = -1
            else:
                # estimating the remaining yield based on the benchmark yield at the time
                yield_4y = self.price_history[-1] / self.price_history[-17]
                ratio = yield_4y / calculation.get_benchmark_4y_yield()
                yeild_missing_year = calculation.get_benchmark_yield()[0] * ratio
                self.yield_5y = yeild_missing_year * yield_4y - 1

            self.yield_5y = calculation.two_point_percentage(self.yield_5y)

        except Exception as e:
            print("FP.init: yeild")
            print(e)
        self.price_history = self.price_history[::2]

        print("FinanceProduct "+self.symbol + " was created")

    def __repr__(self):
        return (
                "\nsymbol: " + str(self.symbol) +
                "\nname: " + str(self.name) +
                "\nprice: " + str(self.current_price) +
                "\ncurrency: " + str(self.currency) +
                "\ntype: " + str(self.product_type) +
                "\nsector: " + str(self.sector) +
                "\nindustry: " + str(self.industry) +
                "\nleveraged: " + str(self.leveraged) +
                #    "\ncountry: " + str(self.country)+
                "\nyield_5y: " + str(self.yield_5y) +
                "\nyield_1y: " + str(self.yield_1y) +
                "\ndiv_per_price: " + str(self.div_per_price) +  # maybe add [-2] ??
                "\nmarket_cap: " + str(self.market_cap) +
                "\navg_volume: " + str(self.avg_volume) +
                "\npe_ratio: " + str(self.pe_ratio) +

                "\nprice_history: " + str(self.price_history) +
                "\nyearly_dividend: " + str(self.yearly_dividend)

        )

    def get_brief_data_show(self):

        return [str(self.symbol),
                str(self.name),
                str(self.current_price),  # maybe combine with currency
                str(self.currency),
                str(self.product_type),
                str(self.sector),
                str(self.industry),
                str(self.leveraged),
                str(self.yield_5y) + '%',
                str(self.yield_1y) + '%',
                str(self.div_per_price[-2]),  # last full year info
                str(calculation.add_prefix(self.market_cap)),
                str(self.profitability) + '%',
                str(self.debt_to_assent) + '%',
                str(self.market_cap) +
                str(self.avg_volume) +
                str(self.pe_ratio) +
                str(self.dates[0]) +
                str(self.price_history) +
                str(self.yearly_dividend)
                ]

    def get_brief_data(self):
        return [self.symbol,
                self.name,
                self.current_price,  # maybe combine with currency
                self.currency,
                self.product_type,
                self.sector,
                self.industry,
                self.leveraged,
                self.yield_5y,
                self.yield_1y,
                self.last_full_year_div,
                self.market_cap,
                self.profitability,
                self.debt_to_assent,
                self.market_cap,
                self.avg_volume,
                self.pe_ratio,
                "/".join(self.dates[0]),  # get the first mesuring date
                self.price_history,
                self.yearly_dividend
                ]

    def get_data_for_xlsx(self):
        data = self.get_brief_data()
        return data[:-2] + list(self.price_history) + list(self.yearly_dividend)


    def get_brief_data_sql(self):

        return [self.symbol,
                self.name,
                self.current_price,
                self.currency,
                (0 if self.product_type == "Stock" else 1),
                calculation.get_sector_index(self.sector),
                self.industry,
                (1 if self.leveraged else 0),
                self.yield_5y,
                self.yield_1y,
                self.div_per_price[-2],
                self.market_cap,
                self.profitability,
                self.debt_to_assent
                ]

    def get_full_data(self):

        """
            getting additinal enformatiom:

            expense_ratio
            top holdings
            top sectors

            all revenue and income
            all assent and debt
            free cash
        """

        # self.country = info["country"] # state

        if self.product_type == "Stock":

            ##### get all income and revenue ######
            url = "https://finance.yahoo.com/quote/" + self.symbol + "/financials?p=" + self.symbol
            html_content = requests.get(url).text  # Make a GET request to fetch the raw HTML content
            soup = BeautifulSoup(html_content, "lxml")  # Parse the html content
            data = soup.find("div", attrs={"class": "D(tbrg)"})

            revenue_data = data.contents[0].contents[0].contents
            revenue = [calculation.convert_string_to_number(x.string) for x in revenue_data[
                                                                               2:]]  # in 0 get the title 1 is not important the rest are the values per year 2019-201
            income_data = data.contents[10].contents[0].contents
            income = [calculation.convert_string_to_number(x.string) for x in income_data[
                                                                              2:]]  # in 0 get the title 1 is not important the rest are the values per year 2019-2016

            ##### get all assent and debt ######
            url = "https://finance.yahoo.com/quote/" + self.symbol + "/balance-sheet?p=" + self.symbol
            html_content = requests.get(url).text  # Make a GET request to fetch the raw HTML content
            soup = BeautifulSoup(html_content, "lxml")  # Parse the html content
            data = soup.find("div", attrs={"class": "D(tbrg)"})

            assent_data = data.contents[0].contents[0].contents
            assent = [calculation.convert_string_to_number(x.string) for x in assent_data[
                                                                              1:]]  # in 0 get the title 1 is not important the rest are the values per year 2019-201
            debt_data = data.contents[11].contents[0].contents
            debt = [calculation.convert_string_to_number(x.string) for x in
                    debt_data[1:]]  # in 0 get the title 1 is not important the rest are the values per year 2019-2016

            ##### get free cash ######
            """"get the stock name and return 2 diamention list with the last 4 years data"""
            url = "https://finance.yahoo.com/quote/" + self.symbol + "/cash-flow?p=" + self.symbol
            html_content = requests.get(url).text  # Make a GET request to fetch the raw HTML content
            soup = BeautifulSoup(html_content, "lxml")  # Parse the html content
            data = soup.find("div", attrs={"class": "D(tbrg)"}).contents[14].contents[0].contents[2:]
            free_cash = [calculation.convert_string_to_number(x.text) for x in data]


        else:  # if product is ETF

            ###### get ETF expense_ratio ######
            url = "https://finance.yahoo.com/quote/" + self.symbol + "/"
            html_content = requests.get(url).text  # Make a GET request to fetch the raw HTML content
            soup = BeautifulSoup(html_content, "lxml")  # Parse the html content
            data = soup.find("div", attrs={"id": "quote-summary"})

            expense_ratio = calculation.two_point_percentage(float(
                data.contents[1].contents[0].contents[0].contents[5].contents[1].text[:-1]))

            ###### get sectors ########
            url = "https://finance.yahoo.com/quote/" + self.symbol + "/holdings?p=" + self.symbol
            html_content = requests.get(url).text  # Make a GET request to fetch the raw HTML content
            soup = BeautifulSoup(html_content, "lxml")  # Parse the html content
            data = soup.find("section", attrs={"class": "Pb(20px) smartphone_Px(20px) smartphone_Mt(20px)"})
            sectors = {}

            sectors["bond share"] = calculation.convert_string_to_number(
                data.contents[0].contents[0].contents[1].contents[1].contents[1].text[:-1])

            if sectors["bond share"] > 0:
                bond_sectors_data = data.contents[1].contents[1].contents[1].contents
                sectors["bond sectors"] = [calculation.convert_string_to_number(x.contents[1].text[:-1]) for x in
                                           bond_sectors_data[1:]]
            else:
                sectors["bond sectors"] = []

            sectors["stocks share"] = calculation.convert_string_to_number(
                data.contents[0].contents[0].contents[1].contents[0].contents[1].text[:-1])

            if sectors["stocks share"] > 0:
                sectors_share_data = data.contents[0].contents[1].contents[1].contents
                sectors["stock sectors"] = [calculation.convert_string_to_number(x.contents[2].text[:-1]) for x in
                                            sectors_share_data[1:]]
            else:
                sectors["stock sectors"] = []

            ###### get holdings ########
            top_holdings_data = data.contents[3].contents[1].contents[1].contents
            top_holdings = [
                [x.contents[0].text, x.contents[1].text, calculation.convert_string_to_number(x.contents[2].text[:-1])]
                for x in top_holdings_data]
            # top holding example [[microsoft corp, MSFT, 11.76], [apple inc, AAPL, 11.09] ... ]

        # TODO update this part

        return [str(self.symbol),
                str(self.name),
                str(self.current_price),
                str(self.currency),
                str(self.product_type),
                str(self.sector),
                str(self.industry),
                str(self.leveraged),
                # str(self.country),
                str(self.price_history),
                str(self.yield_5y),
                str(self.yield_1y),
                str(self.div_per_price),
                str(self.yearly_dividend),
                str(self.market_cap),
                str(self.avg_volume),
                str(self.pe_ratio)]

    def save_as_xlsx(self):
        data = self.get_data_for_xlsx()
        workbook = openpyxl.load_workbook('stock_data.xlsx')
        sheet = workbook.active
        sheet.append(data)
        workbook.save('stock_data.xlsx')

        print("FinanceProduct "+self.symbol + " was saved to excel file")

    @staticmethod
    def clear_xlsx_data():
        path = os.getcwd()
        workbook = xlsxwriter.Workbook(path+'/stock_data.xlsx')
        workbook.add_worksheet('Sheet1')
        workbook = openpyxl.load_workbook('stock_data.xlsx')
        sheet = workbook.active
        sheet.append(list(calculation.headlines))

        workbook.close()


if __name__ == '__main__':
    FinanceProduct.clear_xlsx_data()
    print("satrt")
    fp = FinanceProduct("msft")
    fp.clear_xlsx_data()
    fp.save_as_xlsx()
    fp.save_as_xlsx()

    fp = FinanceProduct("aapl")
    fp.save_as_xlsx()
    fp.save_as_xlsx()

    print("done")
